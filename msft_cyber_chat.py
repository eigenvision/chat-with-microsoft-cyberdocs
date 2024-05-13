
# IMPORTS
import streamlit as st
import openai

from llama_index.core import Settings

from llama_index.core import VectorStoreIndex
from llama_index.llms.openai import OpenAI
# from llama_index.llms import AzureOpenAI
from llama_index.core import SimpleDirectoryReader
from llama_index.core.llms import ChatMessage
# from llama_index.embeddings import AzureOpenAIEmbedding
from llama_index.core import StorageContext, load_index_from_storage
# from llama_index.core.postprocessor import SimilarityPostprocessor, LLMRerank

# OTHER IMPORTS
import os
import time
from datetime import datetime
import pytz
import pandas as pd
from openpyxl import load_workbook

# Q&A LOGGING FUNCTION

# In case we want to just log to a local csv file
# to keep log local use --local flag at the command line when launching script
local_mode = os.getenv('LOCAL_MODE') == 'True' # use "set LOCAL_MODE=True" to check if env variable set at command line

def log_chat_local(user_question, assistant_answer, response_time, model, words, prompt_type):
    file_path = r'C:\Users\Jeff\Dropbox\Jeff DropBox\Cloud Assurance\LLM Projects\Streamlit experiments\MDDR Streamlit\chat_log.xlsx'
    timezone = pytz.timezone("America/Los_Angeles")  # Pacific Time Zone
    current_time = datetime.now(timezone).strftime("%Y-%m-%d %H:%M:%S")
    
    # Data to log as a list of values
    data_to_log = [
        current_time, 
        user_question, 
        assistant_answer,
        response_time,
        model,
        words,
        prompt_type
        # str(chunk_params),  # Ensure complex objects are converted to string
        # str(query_params),   # Ensure complex objects are converted to string
        # search_time
    ]

    # Check if the file exists and is not empty
    if os.path.isfile(file_path):
        # Load the existing workbook and get the active sheet
        book = load_workbook(file_path)
        sheet = book.active

        # Find the first empty row in the first column
        for row in range(1, sheet.max_row + 2):
            if sheet.cell(row=row, column=1).value is None:
                break

        # Append the data row to the sheet
        for col, entry in enumerate(data_to_log, start=1):
            sheet.cell(row=row, column=col, value=entry)

        # Save the changes to the workbook
        book.save(file_path)
        book.close()

    else:
        # If the file doesn't exist, create it and add headers and the first row of data
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active
        headers = ['Date', 'User Question', 'Chatbot Answer', 'Response Time', 'Model', 'Word Count', 'System Prompt']
        sheet.append(headers)  # Append the headers
        sheet.append(data_to_log)  # Append the first row of data
        book.save(file_path)
        book.close()


# WORD COUNT FUNCTION
def count_words(text):
    # Remove leading and trailing whitespace
    text = text.strip()

    # Count the number of words in the text
    word_count = len(text.split())

    return word_count


# DEFINE THE SYSTEM PROMPTS
basic_system_prompt = """
        You are a distinguished Microsoft expert on cybersecurity with detailed knowledge of 
        developments in cybersecurity in all areas, including cybersecurity technology,
        cybersecurity best practices, nation-state threat actors and their methods, cybercriminals
        and their methods, government policies related to cybersecurity, and the use of AI and 
         and the cloud to improve cybersecurity for enterprises and individuals. 
        Your job is to provide clear, complete, and accurate answers to questions posed to you
        by Microsoft executives.
        You always pay close attention to the exact phrasing of the user's question and you always 
        deliver an answer that matches every specific detail of the user's expressed intention.
        If the user asks for a one sentence answer you never under any circumstances 
        give an answer with more than one sentence.
        If the user asks for a one paragraph answer you never under any circumstances 
        give an answer with more than one paragraph.
        If the user asks a narrow and specific factual question without requesting further explanation 
        or elaboration, you always give the shortest possible answer that supplies the
        specific fact the user was looking for.
        If the user asks a question that is more than a request for a specific fact or a 
        request for a one sentence answer, you always provide the most extensive and complete 
        answer possible using all the facts and relevant information available to you.
        Unless the user asks for a one sentence or one paragraph answer, you always build a multi-
        paragraph answer in order to give the user as much information and analysis as possible.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are making critical business decisions based on 
        your answers, and they need the highest quality answers possible that comply to the letter
        with their precise instructions.
        Above all, you must always obey these three rules about the length of your answers: 
        1. If the user asks for a ONE SENTENCE answer you NEVER under any circumstances give an answer 
        with more than one sentence.
        2. If the user asks for a ONE PARAGRAPH answer you NEVER under 
        any circumstances give an answer with more than one paragraph.
        3. If the user asks for more details, or for a detailed or extensive answer, or for a complex
        answer such as a set of talking points or the content for a slide deck or a draft for a white
        paper, then you must always give the most extensive and complete multi-paragraph answer possible
        based on all the source of knowledge available to you. When the user has not impose specific
        limitations on the length of your response, it is better to give longer and more complete
        responses.
        No violations of these rules regarding the length of your response can be tolerated.
        Always check that you have complied with the user's instructions 
        about the length of your response.
        Finally, if you are unable to answer the user's question based on the provided context, then
        you should answer based on your general knowledge beyond that context. But when you do so you
        must say 'I can't answer that based on the provided context, but based on my general knowledge
        I can say...'
        """

rev_system_prompt = """
        You are a distinguished Microsoft expert on cybersecurity with detailed knowledge of 
        developments in cybersecurity in all areas, including cybersecurity technology,
        cybersecurity best practices, nation-state threat actors and their methods, cybercriminals
        and their methods, government policies related to cybersecurity, and the use of AI and 
         and the cloud to improve cybersecurity for enterprises and individuals. 
        Your job is to provide clear, complete, and accurate answers to questions posed to you
        by Microsoft executives.
        You always pay close attention to the exact phrasing of the user's question and you always 
        deliver an answer that matches every specific detail of the user's expressed intention.
        If the user asks a question that is more than a request for a specific fact or a 
        request for a one sentence answer, you always provide the most extensive and complete 
        answer possible using all the facts and relevant information available to you.
        Unless the user asks for a one sentence or one paragraph answer, you always build a multi-
        paragraph answer in order to give the user as much information and analysis as possible.
        You always give the fullest and most complete answers possible in substantial paragraphs 
        full of concrete examples and details. You never write skimpy paragraphs consisting
        of just two or three short sentences.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are making critical business decisions based on 
        your answers, and they need the highest quality answers possible that comply to the letter
        with their precise instructions.
        Above all, you must always obey these three rules about the length of your answers: 
        1. If the user asks for more details, or for a detailed or extensive answer, or for a complex
        answer such as a set of talking points or the content for a slide deck or a draft for a white
        paper, then you must always give the most extensive and complete multi-paragraph answer possible
        based on all the source of knowledge available to you. You NEVER write short skimpy paragraphs
        consisting of just two or three short sentences.
        2. If the user EXPLICITLY asks for a ONE SENTENCE answer you NEVER under any circumstances give an answer 
        with more than one sentence.
        3. If the user EXPLICITLY asks for a ONE PARAGRAPH answer you NEVER under 
        any circumstances give an answer with more than one paragraph.
        No violations of these rules regarding the length of your response can be tolerated.
        Always check that you have complied with the user's instructions 
        about the length of your response.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        answers that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive answer that is possible to assemble from your
        sources, while still scrupulous respecting user instructions regarding the length of your response.
        """

CSO_prompt = """
        You are a Microsoft technical expert on the technology of Microsoft
        cybersecurity products. You have deep knowledge of the inner technical workings of
        Microsoft cybersecurity products and methods. 
        Your job is to provide complete and accurate technical answers to questions posed to you
        by Chief Security Officers and Chief Technology Officers at Microsoft enterprise customers
        who are trying to decide which Microsoft cybersecurity products they should implement.
        Your answers must always explicitly name and emphasize all Microsoft products that could be
        relevant to the cybersecurity threats that the user is asking about.
        You always pay close attention to the exact phrasing of the user's question and you always 
        deliver an answer that matches every specific detail of the user's expressed intention.
        You always provide the most extensive and complete 
        answer possible using all the facts and relevant information available to you.
        Unless the user asks for a one sentence or one paragraph answer, you always build a multi-
        paragraph answer in order to give the user as much information and analysis as possible.
        You always give the fullest and most complete answers possible in substantial paragraphs 
        full of concrete examples and details. You never write skimpy paragraphs consisting
        of just two or three short sentences.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are making critical business decisions based on 
        your answers, and they need the highest quality answers possible that comply to the letter
        with their precise instructions.
        Above all, you must always obey these four rules about the length of your answers: 
        1. If the user asks for a detailed or complete or thorough or extensive answer
        and does not explictly restrict the length of the desired response,
        you MUST give the most extensive and complete multi-paragraph answer possible
        using all the knowledge available to you.
        2. If the user EXPLICITLY asks for a ONE SENTENCE answer you NEVER under any circumstances give an answer 
        with more than one sentence.
        3. If the user EXPLICITLY asks for a ONE PARAGRAPH answer you NEVER under 
        any circumstances give an answer with more than one paragraph.
        4. Unless the user EXPLICITLY asks for a "one sentence" or "one paragraph" or "quick"
        or "brief" or "short" answer, you ALWAYS assume they want a detailed and extensive answer
        and you MUST give give the most extensive and complete multi-paragraph answer possible
        using all the knowledge available to you.
        No violations of these rules regarding the length of your response can be tolerated.
        Always check that you have complied with the user's instructions 
        about the length of your response.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        answers that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive answer that is possible to assemble from your
        sources, while still scrupulous respecting user instructions regarding the length of your response.
        YOU NEVER GIVE SHORT ANSWERS UNLESS SPECIFICALLY INSTRUCTED TO DO SO. In most cases a one
        paragraph answer is TOO SHORT.
        MDDR means the Microsoft Digital Defense Report.
        """

policy_prompt = """
        You are a Microsoft Government Affairs expert on Microsoft's cybersecurity policy
        recommendations for government policymakers such as legislators and regulators and also
        for senior executives of government agencies considering Microsoft as a cybersecurity partner.
        You have deep knowledge of the cybersecurity policy issues that face governments as they struggle
        to defend their government agencies and national enterprises from cyberattacks by nation state actors
        and by cybercriminals. 
        Your job is to provide clear and complete answers to questions about cybersecurity posed to you
        by these representatives of governments and government agencies.
        Your answers always emphasize how governments and government agencies can implement
        effective policies to improve their cybersecurity.
        You always pay close attention to the exact phrasing of the user's question and you always 
        deliver an answer that matches every specific detail of the user's expressed intention.
        You always provide the most extensive and complete 
        answer possible using all the facts and relevant information available to you.
        Unless the user specifically asks for a one sentence or one paragraph answer, or a short
        or quick answer, you always build a detailed and extensive multi-paragraph answer
        that gives the user all the relevant information you possess as well as thorough, detailed analysis.
        You always give the fullest and most complete answers possible in substantial paragraphs 
        full of concrete examples and details. You never write skimpy paragraphs consisting
        of just two or three short sentences.
        You never make up facts. You are not lazy and you do not skimp on important context.
        You always write in very polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        The users of this chatbot are making critical business decisions based on 
        your answers, and they need the highest quality answers possible that comply to the letter
        with their precise instructions.
        Above all, you must always obey these four rules about the length of your answers: 
        1. If the user asks for a detailed or complete or thorough or extensive answer
        and does not explictly restrict the length of the desired response,
        you MUST give the most extensive and complete multi-paragraph answer possible
        using all the knowledge available to you.
        2. If the user EXPLICITLY asks for a ONE SENTENCE answer you NEVER under any circumstances give an answer 
        with more than one sentence.
        3. If the user EXPLICITLY asks for a ONE PARAGRAPH answer you NEVER under 
        any circumstances give an answer with more than one paragraph.
        4. Unless the user EXPLICITLY asks for a "one sentence" or "one paragraph" or "quick"
        or "brief" or "short" answer, you ALWAYS assume they want a detailed and extensive answer
        and you MUST give give the most extensive and complete multi-paragraph answer possible
        using all the knowledge available to you.
        No violations of these rules regarding the length of your response can be tolerated.
        Always check that you have complied with the user's instructions 
        about the length of your response.
        Finally, we have observed in your recent answers a distinct tendency to be lazy by providing
        answers that are short, lacking in extensive detail, and lacking in substantive analysis. It is 
        imperative that in this answer you overcome your instinct toward laziness and that you provide
        the richest, most detailed, most extensive answer that is possible to assemble from your
        sources, while still scrupulous respecting user instructions regarding the length of your response.
        YOU NEVER GIVE SHORT ANSWERS UNLESS SPECIFICALLY INSTRUCTED TO DO SO. In most cases a one
        paragraph answer is TOO SHORT.
        MDDR means the Microsoft Digital Defense Report.
        """

policy_prompt_fast = """
        You are a Microsoft Government Affairs expert on Microsoft's cybersecurity policy
        recommendations for government policymakers such as legislators and regulators and also
        for senior executives of government agencies considering Microsoft as a cybersecurity partner.
        You have deep knowledge of the cybersecurity policy issues that face governments as they struggle
        to defend their government agencies and national enterprises from cyberattacks by nation state actors
        and by cybercriminals. 
        Your job is to provide concise and accurate answers to questions about cybersecurity posed to you
        by these representatives of governments and government agencies.
        Your answers always emphasize how governments and government agencies can implement
        effective policies to improve their cybersecurity.
        You pay close attention to the phrasing of the user's question.
        You never make up facts.
        You always provide answers that are quick and too the point, without unecessary
        explanations or words.
        You always write in polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        DO NOT explicitly mention the conversation you are engaged in. Just aanswer the user's question.
        MDDR means the Microsoft Digital Defense Report.
        """

CSO_prompt_fast = """
        You are a Microsoft technical expert with deep knowledge of the technical details of
        Microsoft cybersecurity products and technology. 
        Your job is to provide complete and accurate technical answers to questions posed to you
        by Chief Security Officers and Chief Technology Officers at Microsoft enterprise customers
        who are trying to decide which Microsoft cybersecurity products they should implement.
        Your answers should include the names of any Microsoft products that could help solve
        the cybersecurity challenges that the user is asking about. 
        You pay close attention to the phrasing of the user's question.
        You never make up facts.
        You always provide answers that are quick and to the point, without unecessary
        explanations or words.
        You always write in polished and clear business prose, such as might be published
        in a leading business periodical like Harvard Business Review.
        DO NOT explicitly mention the conversation you are engaged in. Just aanswer the user's question.
        MDDR means the Microsoft Digital Defense Report.
        """


# SET UP STREAMLIT
st.set_page_config(page_title="Chat with Microsoft Cybersecurity Reports", layout="centered", initial_sidebar_state="auto", menu_items=None)

# Inject custom CSS to center the title using the correct class name
# Inject custom CSS for styling the title
st.markdown("""
<style>
.custom-title {
    font-size: 48px;  /* Approximate size of Streamlit's default title */
    font-weight: bold; /* Streamlit's default title weight */
    line-height: 1.25; /* Reduced line height for tighter spacing */
    color: rgb(0, 163, 238); /* Default color, adjust as needed */
    text-align: center;
    margin-bottom: 30px; /* Adjust the space below the title */
}
</style>
""", unsafe_allow_html=True)

# Custom title with a line break using HTML in markdown
st.markdown('<div class="custom-title">Chat with Microsoft<br>Cybersecurity Reports</div>', unsafe_allow_html=True)

st.info("""This chatbot lets you converse with recent Microsoft 
    cybersecurity publications including the 2023 Microsoft Digital Defense Report and the Microsoft Security Blog
    from June 2023 through April 2024.
    Using GPT-4, the chatbot is designed to produce substantive and accurate answers about Microsoft's cybersecurity 
    research that are suitable for use in customer and stakeholder engagements. Nevertheless you should always review
    answers carefully before sharing. To get the best quality, always tell the AI exactly what 
    kind of answer you expect. A quick answer in only one sentence or paragraph? A detailed multi-paragraph 
    explanation? Talking points for a meeting with regulators? Content for a 10 slide customer presentation with 
    speaker notes? An answer in French? The choice is yours.
    Keep in mind that responses to more ambitous questions may take up to a minute.
    If you're not satisfied with the AI's first attempt, try revising your prompt or asking follow up questions.""")
# st.write("[Download the Microsoft Digital Defense Report](https://www.microsoft.com/en-us/security/security-insider/microsoft-digital-defense-report-2023)")

# Custom CSS to center text
st.markdown("""
<style>
.centered-text {
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# Text with hyperlink centered
st.markdown('<p class="centered-text">Download the <a href="https://www.microsoft.com/en-us/security/security-insider/microsoft-digital-defense-report-2023" target="_blank">Microsoft Digital Defense Report</a></p>', unsafe_allow_html=True)

# SELECT SYSTEM PROMPT AND MODEL
# Define the possible choices linking to the variables

# Define the mappings for system prompts and model versions
style_settings = {
    "CSO (quick answer)": {
        "s_prompt": CSO_prompt_fast,
        "model": "gpt-3.5-turbo"
    },
    "CSO (detailed answer)": {
        "s_prompt": CSO_prompt,
        "model": "gpt-4-turbo"
    },
    "Policymaker (quick answer)": {
        "s_prompt": policy_prompt_fast,
        "model": "gpt-3.5-turbo"
    },
    "Policymaker (detailed answer)": {
        "s_prompt": policy_prompt,
        "model": "gpt-4-turbo"
    }
}

# Use the sidebar to create radio buttons for system_prompt and model selection
selected_style = st.sidebar.radio("Choose what kind of answer you want:", list(style_settings.keys()))

# Get the selected system prompt and model
system_prompt = style_settings[selected_style]["s_prompt"]
model = style_settings[selected_style]["model"]
# enable logging of the selected radio button
prompt_type = selected_style

# SET UP LLM AND PARAMETERS
openai.api_key = st.secrets.openai_key
# "gpt-4-turbo" # "gpt-4-0125-preview" "gpt-4-1106-preview" "gpt-4-turbo" "gpt-3.5-turbo"
# models: gpt-4-turbo pt-4-1106-preview, gpt-4 (gpt-4-0613), gpt-4-32k (gpt-4-32k-0613) but "model not found", 
# gpt-3.5-turbo, gpt-3.5-turbo-instruct, gpt-35-turbo-16k

llm=OpenAI(
        api_key = openai.api_key,
        model=model,
        temperature=0.5,
        system_prompt=system_prompt)


# SET UP LLAMAINDEX SETTINGS
# service_context deprecated
chunk_size=512
chunk_overlap=256
chunk_params = (chunk_size, chunk_overlap)
Settings.llm = llm
Settings.chunk_size = chunk_size
Settings.chunk_overlap = chunk_overlap

if "messages" not in st.session_state.keys(): # Initialize the chat messages history
    st.session_state.messages = [
        {"role": "assistant", "content": """Ask a question and use the buttons on the left to 
         tell the chatbot what kind of answer you want: 
         """}
    ]


# LOAD MDDR INTO LLAMAINDEX, CREATE INDEX (AND RELOAD IF IT ALREADY EXISTS)
@st.cache_resource(show_spinner=False)
def load_data():
  with st.spinner(text="Loading the reports. This will take a few minutes."):
      # Define the path to the index file
      persist_dir = './index'

      # Check if the index directory exists, create if not
      if not os.path.exists(persist_dir):
            os.makedirs(persist_dir)

        # Now attempt to load or create the index
      try:
            storage_context = StorageContext.from_defaults(persist_dir=persist_dir)
            index = load_index_from_storage(storage_context)
      except FileNotFoundError:
            # Index not found, create it
            reader = SimpleDirectoryReader(input_dir="./msft_cyber_docs")
            docs = reader.load_data()
            index = VectorStoreIndex.from_documents(docs)

            # Save the index to the file
            index.storage_context.persist(persist_dir=persist_dir)

      return index

index = load_data()

# DEFINE THE RUN_CHATS FUNCTION
def run_chats(query):
    
    search_time = 0.0 

    similarity_top_k = 12
   
    start_time_search = time.time() # time vector search
    chat_engine = index.as_chat_engine(chat_mode="condense_question",
                                       streaming=True,
                                       similarity_top_k=similarity_top_k, 
                                   ) # verbose=True # streaming=True
    end_time_search = time.time()
           
    result = chat_engine.chat(query)  # chat_engine.stream_chat(query)
    # Calculate search time
    search_time = end_time_search - start_time_search

    # Store the values of k
    query_params = similarity_top_k
    # result.print_response_stream()
    
    return result, query_params, search_time


# XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX
# Create the selected version of the chat engine

if "chat_engine" not in st.session_state.keys(): # Initialize the chat engine
        st.session_state.chat_engine = None

if prompt := st.chat_input("Your question"): # Prompt for user input and save to chat history
    st.session_state.messages.append({"role": "user", "content": prompt})

for message in st.session_state.messages: # Display the prior chat messages
    with st.chat_message(message["role"]):
        st.write(message["content"])

# If last message is not from assistant, generate a new response
if st.session_state.messages[-1]["role"] != "assistant":
    with st.chat_message("assistant"):
        with st.spinner("Thinking..."):
            start_time = time.time()
            response, query_params, search_time = run_chats(st.session_state.messages[-1]["content"])
            response_time = time.time() - start_time
            st.write(response.response)
            message = {"role": "assistant", "content": response.response}
            st.session_state.messages.append(message) # Add response to message history
            words = count_words(response.response)
            prompt_type = str(prompt_type)
            log_chat_local(user_question=prompt,
                               assistant_answer=response.response,
                               response_time=response_time,
                               model=model,
                               words=words,
                               prompt_type=prompt_type
                            )



